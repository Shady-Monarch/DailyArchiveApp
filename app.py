import io
import json
from datetime import date, datetime, timedelta
from pathlib import Path

import gspread
import pandas as pd
import streamlit as st
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import Flow
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

APP_DIR = Path(__file__).resolve().parent
TOKEN_PATH = APP_DIR / ".streamlit" / "google_token.json"
OAUTH_PKCE_PATH = APP_DIR / ".streamlit" / "oauth_pkce.json"
SHEET_NAME = "Daily Log"
# Spreadsheets-only scope — enough to read/write a sheet you already own/share.
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

TEMPLATE = {
    "DEV – Code / Bug Fixes": ["JIRA ID", "JIRA Description", "MR Link"],
    "DEV – POC": ["JIRA ID", "JIRA Description", "Confluence Link"],
    "DEV – MR Review": ["MR Link", "No. of Comments"],
    "Meetings": ["Purpose / Topic Discussed", "Meeting Type"],
    "Miscellaneous": ["Miscellaneous Tasks"],
}

COLUMNS = ["Date", "Category", "Field", "Details", "Created At", "Updated At"]

st.set_page_config(page_title="Daily Work Archive", page_icon="🗂️", layout="wide")


def empty_df():
    return pd.DataFrame(columns=COLUMNS)


def require_secrets():
    missing = []
    oauth = st.secrets.get("google_oauth", {})
    if not oauth.get("client_id") or not oauth.get("client_secret"):
        missing.append("google_oauth.client_id / google_oauth.client_secret")
    if "spreadsheet_id" not in st.secrets and "spreadsheet_url" not in st.secrets:
        missing.append("spreadsheet_id (or spreadsheet_url)")
    if missing:
        st.error(
            "Google Sheets OAuth is not configured. Add these to `.streamlit/secrets.toml` "
            "(local) or Streamlit Cloud secrets:\n\n- "
            + "\n- ".join(missing)
        )
        st.info("See README.md for OAuth setup steps (no service account key required).")
        st.stop()


def detect_redirect_uri_from_request():
    """Build redirect URI from the current browser host (works on Cloud + local)."""
    try:
        headers = st.context.headers
    except Exception:
        return None

    host = headers.get("X-Forwarded-Host") or headers.get("Host")
    if not host:
        return None

    host = host.split(",")[0].strip()
    proto = headers.get("X-Forwarded-Proto")
    if proto:
        proto = proto.split(",")[0].strip()
    elif "streamlit.app" in host or "streamlitapp.com" in host:
        proto = "https"
    else:
        proto = "http"

    if host.endswith(":443") and proto == "https":
        host = host[:-4]
    elif host.endswith(":80") and proto == "http":
        host = host[:-3]

    return f"{proto}://{host}"


def get_redirect_uri():
    # Prefer the live request host so Cloud doesn't accidentally keep using localhost
    # from secrets. Fall back to secrets, then local default.
    detected = detect_redirect_uri_from_request()
    if detected:
        return detected.rstrip("/")

    oauth = st.secrets.get("google_oauth", {})
    if oauth.get("redirect_uri"):
        return str(oauth["redirect_uri"]).rstrip("/")

    return "http://localhost:8501"


def oauth_client_config():
    oauth = st.secrets["google_oauth"]
    return {
        "web": {
            "client_id": oauth["client_id"],
            "client_secret": oauth["client_secret"],
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
            "redirect_uris": [get_redirect_uri()],
        }
    }


def credentials_to_dict(creds):
    return {
        "token": creds.token,
        "refresh_token": creds.refresh_token,
        "token_uri": creds.token_uri,
        "client_id": creds.client_id,
        "client_secret": creds.client_secret,
        "scopes": list(creds.scopes or SCOPES),
    }


def save_token(creds):
    TOKEN_PATH.parent.mkdir(parents=True, exist_ok=True)
    TOKEN_PATH.write_text(json.dumps(credentials_to_dict(creds), indent=2))


def load_token_file():
    if not TOKEN_PATH.exists():
        return None
    try:
        return json.loads(TOKEN_PATH.read_text())
    except Exception:
        return None


def load_pkce_store():
    if not OAUTH_PKCE_PATH.exists():
        return {}
    try:
        return json.loads(OAUTH_PKCE_PATH.read_text())
    except Exception:
        return {}


def save_pkce_store(store):
    OAUTH_PKCE_PATH.parent.mkdir(parents=True, exist_ok=True)
    OAUTH_PKCE_PATH.write_text(json.dumps(store, indent=2))


def remember_pkce(state, code_verifier, auth_url, redirect_uri):
    store = load_pkce_store()
    store[state] = {
        "code_verifier": code_verifier,
        "auth_url": auth_url,
        "redirect_uri": redirect_uri,
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    save_pkce_store(store)


def pop_pkce(state):
    store = load_pkce_store()
    entry = store.pop(state, None)
    save_pkce_store(store)
    return entry


def clear_login():
    st.session_state.pop("google_token", None)
    if TOKEN_PATH.exists():
        TOKEN_PATH.unlink()


def build_credentials(info):
    if not info:
        return None
    creds = Credentials.from_authorized_user_info(info, SCOPES)
    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())
        st.session_state["google_token"] = credentials_to_dict(creds)
        save_token(creds)
    return creds


def get_credentials():
    # 1) In-memory session token
    if "google_token" in st.session_state:
        creds = build_credentials(st.session_state["google_token"])
        if creds and creds.valid:
            return creds

    # 2) Optional refresh token stored in Streamlit secrets (useful on Cloud)
    oauth = st.secrets["google_oauth"]
    if oauth.get("refresh_token"):
        info = {
            "token": None,
            "refresh_token": oauth["refresh_token"],
            "token_uri": "https://oauth2.googleapis.com/token",
            "client_id": oauth["client_id"],
            "client_secret": oauth["client_secret"],
            "scopes": SCOPES,
        }
        creds = build_credentials(info)
        if creds and creds.valid:
            st.session_state["google_token"] = credentials_to_dict(creds)
            return creds

    # 3) Local token file from a previous login
    info = load_token_file()
    if info:
        creds = build_credentials(info)
        if creds and creds.valid:
            st.session_state["google_token"] = credentials_to_dict(creds)
            return creds

    return None


def make_oauth_flow(*, code_verifier=None, autogenerate_code_verifier=True):
    return Flow.from_client_config(
        oauth_client_config(),
        scopes=SCOPES,
        code_verifier=code_verifier,
        autogenerate_code_verifier=autogenerate_code_verifier,
    )


def exchange_code_for_credentials(code, state):
    # Streamlit often loses session_state on the Google redirect, so PKCE lives on disk.
    entry = pop_pkce(state) if state else None
    code_verifier = None
    if entry:
        code_verifier = entry.get("code_verifier")
    if not code_verifier:
        code_verifier = st.session_state.get("oauth_code_verifier")
    if not code_verifier:
        raise RuntimeError(
            "Sign-in could not find the temporary login key. "
            "Click Reset sign-in link, then Sign in with Google again."
        )

    flow = make_oauth_flow(
        code_verifier=code_verifier,
        autogenerate_code_verifier=False,
    )
    flow.redirect_uri = get_redirect_uri()
    flow.fetch_token(code=code)
    for key in ("oauth_code_verifier", "oauth_state", "oauth_auth_url"):
        st.session_state.pop(key, None)
    return flow.credentials


def begin_oauth_login():
    redirect_uri = get_redirect_uri()

    # Reuse one pending login only if it was minted for this exact redirect URI.
    state = st.session_state.get("oauth_state")
    store = load_pkce_store()
    entry = store.get(state) if state else None
    if (
        state
        and entry
        and entry.get("auth_url")
        and entry.get("redirect_uri") == redirect_uri
    ):
        auth_url = entry["auth_url"]
    else:
        if state:
            pop_pkce(state)
        flow = make_oauth_flow(autogenerate_code_verifier=True)
        flow.redirect_uri = redirect_uri
        auth_url, state = flow.authorization_url(
            access_type="offline",
            prompt="consent",
        )
        remember_pkce(state, flow.code_verifier, auth_url, redirect_uri)
        st.session_state["oauth_auth_url"] = auth_url
        st.session_state["oauth_state"] = state
        st.session_state["oauth_code_verifier"] = flow.code_verifier

    st.link_button("🔐 Sign in with Google", auth_url, type="primary", use_container_width=True)
    st.warning(
        f"Add this **exact** URI in Google Cloud → OAuth client → Authorized redirect URIs "
        f"(and JS origins):\n\n`{redirect_uri}`"
    )
    st.info("Click once, finish Google consent, and wait to be redirected back here.")
    if st.button("Reset sign-in link"):
        if st.session_state.get("oauth_state"):
            pop_pkce(st.session_state["oauth_state"])
        for key in ("oauth_code_verifier", "oauth_state", "oauth_auth_url"):
            st.session_state.pop(key, None)
        st.rerun()


def ensure_google_login():
    """Handle OAuth callback and block the app until the user is signed in."""
    params = st.query_params

    if "code" in params:
        try:
            state = params.get("state")
            creds = exchange_code_for_credentials(params["code"], state)
            st.session_state["google_token"] = credentials_to_dict(creds)
            save_token(creds)
            st.query_params.clear()
            st.success("Signed in with Google.")
            st.rerun()
        except Exception as exc:
            st.query_params.clear()
            st.error(f"Google sign-in failed: {exc}")
            begin_oauth_login()
            st.stop()

    if "error" in params:
        st.error(f"Google sign-in error: {params.get('error')}")
        st.query_params.clear()
        st.stop()

    creds = get_credentials()
    if creds and creds.valid:
        return creds

    st.title("🗂️ Daily Work Archive")
    st.warning("Sign in with your Google account to access your archive Sheet.")
    begin_oauth_login()
    st.stop()


def get_gspread_client():
    creds = get_credentials()
    if not creds or not creds.valid:
        st.error("Google credentials expired. Please sign in again.")
        clear_login()
        st.rerun()
    return gspread.authorize(creds)


def get_worksheet():
    client = get_gspread_client()
    if "spreadsheet_id" in st.secrets:
        spreadsheet = client.open_by_key(st.secrets["spreadsheet_id"])
    else:
        spreadsheet = client.open_by_url(st.secrets["spreadsheet_url"])

    try:
        worksheet = spreadsheet.worksheet(SHEET_NAME)
    except gspread.WorksheetNotFound:
        worksheet = spreadsheet.add_worksheet(title=SHEET_NAME, rows=1000, cols=len(COLUMNS))
        worksheet.append_row(COLUMNS)
    else:
        values = worksheet.get_all_values()
        if not values:
            worksheet.append_row(COLUMNS)
    return worksheet


def normalize_dataframe(df):
    if df is None or df.empty:
        return empty_df()

    for col in COLUMNS:
        if col not in df.columns:
            df[col] = ""

    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date
    for col in ["Category", "Field", "Details", "Created At", "Updated At"]:
        df[col] = df[col].fillna("").astype(str)

    return df[COLUMNS].dropna(subset=["Date"]).reset_index(drop=True)


def load_data():
    worksheet = get_worksheet()
    try:
        records = worksheet.get_all_records()
    except Exception as exc:
        st.error(f"Failed to read Google Sheet: {exc}")
        return empty_df()

    if not records:
        return empty_df()

    return normalize_dataframe(pd.DataFrame(records))


def rows_for_sheet(df):
    if df.empty:
        return [COLUMNS]

    out = df.copy()
    out["Date"] = out["Date"].astype(str)
    for col in ["Category", "Field", "Details", "Created At", "Updated At"]:
        out[col] = out[col].fillna("").astype(str)
    return [COLUMNS] + out[COLUMNS].values.tolist()


def save_dataframe(df):
    worksheet = get_worksheet()
    payload = rows_for_sheet(df)
    worksheet.clear()
    worksheet.update(payload, value_input_option="USER_ENTERED")


def upsert_day(day, values, existing):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    df = existing.copy()

    # Remove all template rows for this date. The submitted form becomes the source of truth.
    mask = df["Date"] == day
    old = df[~mask].copy()

    old_day = df[mask].copy()
    old_created = {}
    for _, r in old_day.iterrows():
        old_created[(r["Category"], r["Field"])] = r["Created At"]

    rows = []
    for category, fields in TEMPLATE.items():
        for field in fields:
            value = values.get((category, field), "").strip()
            if not value:
                continue
            rows.append({
                "Date": day,
                "Category": category,
                "Field": field,
                "Details": value,
                "Created At": old_created.get((category, field), now),
                "Updated At": now,
            })

    new_day = pd.DataFrame(rows, columns=COLUMNS)
    result = pd.concat([old, new_day], ignore_index=True)
    if not result.empty:
        result["Date"] = pd.to_datetime(result["Date"]).dt.date
    save_dataframe(result)
    return len(rows)


def delete_day(day):
    df = load_data()
    save_dataframe(df[df["Date"] != day].reset_index(drop=True))


def move_day(from_day, to_day, existing):
    """Move all entries from from_day to to_day. Target day must be empty."""
    if from_day == to_day:
        raise ValueError("Source and target dates must be different.")

    df = existing.copy()
    source = df[df["Date"] == from_day]
    if source.empty:
        raise ValueError(f"No entries found for {from_day.strftime('%d %b %Y')}.")

    if not df[df["Date"] == to_day].empty:
        raise ValueError(
            f"Cannot move: {to_day.strftime('%d %b %Y')} already has entries. "
            "Edit or delete that day first."
        )

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    moved = source.copy()
    moved["Date"] = to_day
    moved["Updated At"] = now

    result = pd.concat([df[df["Date"] != from_day], moved], ignore_index=True)
    result["Date"] = pd.to_datetime(result["Date"]).dt.date
    save_dataframe(result)
    return len(moved)


def filtered_data(df, start, end, category=None, keyword=""):
    out = df[(df["Date"] >= start) & (df["Date"] <= end)].copy()
    if category and category != "All":
        out = out[out["Category"] == category]
    if keyword.strip():
        q = keyword.strip().lower()
        mask = (
            out["Details"].str.lower().str.contains(q, na=False)
            | out["Field"].str.lower().str.contains(q, na=False)
            | out["Category"].str.lower().str.contains(q, na=False)
        )
        out = out[mask]
    return out


def make_report(df, start, end, style="Detailed"):
    data = df[(df["Date"] >= start) & (df["Date"] <= end)].copy()
    if data.empty:
        return "No work was recorded for the selected period."

    title = f"# Work Report: {start.strftime('%d %b %Y')} to {end.strftime('%d %b %Y')}"
    lines = [title, ""]

    if style == "Executive":
        categories = data.groupby("Category").size().sort_values(ascending=False)
        lines += ["## Executive Summary", ""]
        for category, count in categories.items():
            lines.append(f"- **{category}:** {count} recorded items")
        lines += ["", "## Highlights", ""]
        for _, r in data.iterrows():
            details = str(r["Details"]).strip()
            if details and details.lower() != "nan":
                lines.append(f"- **{r['Category']} — {r['Field']}:** {details}")
        return "\n".join(lines)

    for d in sorted(data["Date"].unique()):
        lines.append(f"## {pd.Timestamp(d).strftime('%A, %d %b %Y')}")
        day = data[data["Date"] == d]
        for category in TEMPLATE:
            group = day[day["Category"] == category]
            if group.empty:
                continue
            lines.append(f"### {category}")
            for _, r in group.iterrows():
                details = str(r["Details"]).strip()
                if details and details.lower() != "nan":
                    lines.append(f"- **{r['Field']}:** {details}")
            lines.append("")
    return "\n".join(lines)


def style_sheet(ws):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    widths = [14, 28, 32, 90, 22, 22]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    if ws.max_row >= 1:
        ws.auto_filter.ref = ws.dimensions


def excel_bytes(df):
    """Build a downloadable Excel snapshot from the current archive."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(COLUMNS)

    for row in df.itertuples(index=False):
        values = list(row)
        values[0] = str(values[0])
        ws.append(values)

    style_sheet(ws)

    for row in range(2, ws.max_row + 1):
        field = str(ws.cell(row, 3).value or "")
        details = str(ws.cell(row, 4).value or "")
        if field in {"JIRA ID", "MR Link", "Confluence Link"}:
            if details.startswith(("http://", "https://")):
                ws.cell(row, 4).hyperlink = details
                ws.cell(row, 4).font = Font(color="0563C1", underline="single")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def has_any_filled_field(values):
    return any(str(v).strip() for v in values.values())


# ---------------- UI ----------------
require_secrets()
ensure_google_login()

top = st.columns([6, 1])
with top[0]:
    st.title("🗂️ Daily Work Archive")
    st.caption("A searchable, editable work journal backed by Google Sheets (your Google login).")
with top[1]:
    if st.button("Log out"):
        clear_login()
        st.rerun()

data = load_data()

# Show flash messages set before st.rerun() so the user actually sees them.
if "flash_success" in st.session_state:
    st.success(st.session_state.pop("flash_success"))
if "flash_error" in st.session_state:
    st.error(st.session_state.pop("flash_error"))

NAV_ENTRY = "➕ Daily Entry"
NAV_SEARCH = "🔎 Search & Edit"
NAV_REPORTS = "📊 Reports"
NAV_ARCHIVE = "🗃️ Archive"

# Radio keeps the selected section across reruns; st.tabs always resets to the first tab.
section = st.radio(
    "Section",
    [NAV_ENTRY, NAV_SEARCH, NAV_REPORTS, NAV_ARCHIVE],
    horizontal=True,
    label_visibility="collapsed",
    key="main_nav",
)

# DAILY ENTRY
if section == NAV_ENTRY:
    selected_date = st.date_input("Work date", value=date.today(), key="entry_date")
    existing_day = data[data["Date"] == selected_date]
    day_already_exists = not existing_day.empty

    def existing_value(category, field):
        hit = existing_day[
            (existing_day["Category"] == category) & (existing_day["Field"] == field)
        ]
        return "" if hit.empty else str(hit.iloc[-1]["Details"])

    if day_already_exists:
        st.error(
            f"An entry already exists for {selected_date.strftime('%d %b %Y')}. "
            "Use **Search & Edit** to update it, or delete it there before creating a new entry for this day."
        )
    else:
        st.info("Fill in the fields below and save. Blank fields are skipped.")

    form_values = {}
    for category, fields in TEMPLATE.items():
        with st.expander(category, expanded=not day_already_exists):
            for field in fields:
                form_values[(category, field)] = st.text_input(
                    field,
                    value=existing_value(category, field),
                    key=f"entry::{selected_date}::{category}::{field}",
                    disabled=day_already_exists,
                )

    if st.button(
        "💾 Save Day",
        type="primary",
        use_container_width=True,
        disabled=day_already_exists,
    ):
        if not data[data["Date"] == selected_date].empty:
            st.session_state["flash_error"] = (
                f"Cannot save: an entry already exists for {selected_date.strftime('%d %b %Y')}. "
                "Edit or delete it from Search & Edit first."
            )
            st.rerun()
        if not has_any_filled_field(form_values):
            st.session_state["flash_error"] = (
                "Cannot save: fill in at least one field before saving."
            )
            st.rerun()
        count = upsert_day(selected_date, form_values, data)
        st.session_state["flash_success"] = (
            f"Saved {count} non-empty entries for {selected_date.strftime('%d %b %Y')}."
        )
        st.rerun()


# SEARCH & EDIT
if section == NAV_SEARCH:
    st.subheader("Find past work")

    c1, c2, c3 = st.columns([1, 1, 1.5])
    s_from = c1.date_input("From", value=date.today() - timedelta(days=30), key="search_from")
    s_to = c2.date_input("To", value=date.today(), key="search_to")
    s_category = c3.selectbox("Category", ["All"] + list(TEMPLATE.keys()), key="search_category")
    keyword = st.text_input("Keyword / JIRA / MR / meeting topic", key="search_keyword")

    if s_from > s_to:
        st.error("From date cannot be after To date.")
    else:
        results = filtered_data(data, s_from, s_to, s_category, keyword)
        st.metric("Matching entries", len(results))

        if results.empty:
            st.info("No matching entries.")
        else:
            st.dataframe(
                results[["Date", "Category", "Field", "Details", "Updated At"]],
                use_container_width=True,
                hide_index=True,
            )

            st.divider()
            st.subheader("Edit a day")

            days = sorted(results["Date"].unique(), reverse=True)
            edit_day = st.selectbox(
                "Choose a date to edit",
                days,
                format_func=lambda d: pd.Timestamp(d).strftime("%A, %d %b %Y"),
            )

            day_rows = data[data["Date"] == edit_day]

            def day_value(category, field):
                hit = day_rows[
                    (day_rows["Category"] == category) & (day_rows["Field"] == field)
                ]
                return "" if hit.empty else str(hit.iloc[-1]["Details"])

            edit_values = {}
            for category, fields in TEMPLATE.items():
                with st.expander(category):
                    for field in fields:
                        edit_values[(category, field)] = st.text_input(
                            field,
                            value=day_value(category, field),
                            key=f"edit::{edit_day}::{category}::{field}",
                        )

            ec1, ec2 = st.columns(2)
            with ec1:
                if st.button("💾 Update Selected Day", type="primary", use_container_width=True):
                    if not has_any_filled_field(edit_values):
                        st.session_state["flash_error"] = (
                            "Cannot update: fill in at least one field, or delete the day instead."
                        )
                        st.rerun()
                    count = upsert_day(edit_day, edit_values, data)
                    st.session_state["flash_success"] = f"Updated {count} entries."
                    st.rerun()
            with ec2:
                if st.button("🗑️ Delete Entire Day", use_container_width=True):
                    st.session_state["confirm_delete"] = str(edit_day)

            if st.session_state.get("confirm_delete") == str(edit_day):
                st.warning(f"Delete all archived entries for {pd.Timestamp(edit_day).strftime('%d %b %Y')}?")
                d1, d2 = st.columns(2)
                with d1:
                    if st.button("Yes, permanently delete", type="primary"):
                        delete_day(edit_day)
                        st.session_state.pop("confirm_delete", None)
                        st.session_state["flash_success"] = "Day deleted."
                        st.rerun()
                with d2:
                    if st.button("Cancel"):
                        st.session_state.pop("confirm_delete", None)
                        st.rerun()

    st.divider()
    st.subheader("Move a day")
    st.caption(
        "Change the date of every entry on a day to a different date. "
        "The target date must have no entries."
    )

    if data.empty:
        st.info("No days available to move.")
    else:
        recorded_days = sorted(data["Date"].unique(), reverse=True)
        mc1, mc2 = st.columns(2)
        move_from = mc1.selectbox(
            "Move entries from",
            recorded_days,
            format_func=lambda d: pd.Timestamp(d).strftime("%A, %d %b %Y"),
            key="move_from_day",
        )
        move_to = mc2.date_input(
            "Move entries to",
            value=date.today(),
            key="move_to_day",
        )

        source_count = int((data["Date"] == move_from).sum())
        target_occupied = not data[data["Date"] == move_to].empty
        st.write(f"Source has **{source_count}** entries.")

        if move_from == move_to:
            st.warning("Pick a different target date.")
        elif target_occupied:
            st.error(
                f"{move_to.strftime('%d %b %Y')} already has entries. "
                "Delete or edit that day first before moving here."
            )
        else:
            if st.button("📅 Move day", type="primary", use_container_width=True):
                try:
                    count = move_day(move_from, move_to, data)
                    st.session_state["flash_success"] = (
                        f"Moved {count} entries from {move_from.strftime('%d %b %Y')} "
                        f"to {move_to.strftime('%d %b %Y')}."
                    )
                except ValueError as exc:
                    st.session_state["flash_error"] = str(exc)
                st.rerun()


# REPORTS
if section == NAV_REPORTS:
    st.subheader("Generate a report")

    quick = st.selectbox(
        "Quick range",
        ["Custom", "This week", "Last week", "This month", "Last month"],
        key="quick_range",
    )

    today = date.today()
    if quick == "This week":
        report_from = today - timedelta(days=today.weekday())
        report_to = today
    elif quick == "Last week":
        report_to = today - timedelta(days=today.weekday() + 1)
        report_from = report_to - timedelta(days=6)
    elif quick == "This month":
        report_from = today.replace(day=1)
        report_to = today
    elif quick == "Last month":
        first_this = today.replace(day=1)
        report_to = first_this - timedelta(days=1)
        report_from = report_to.replace(day=1)
    else:
        report_from = st.date_input("From", value=today - timedelta(days=6), key="report_from_custom")
        report_to = st.date_input("To", value=today, key="report_to_custom")

    report_style = st.radio(
        "Report style",
        ["Detailed", "Executive"],
        horizontal=True,
        help="Executive is shorter and better suited for status updates.",
    )

    if report_from > report_to:
        st.error("From date cannot be after To date.")
    else:
        report_data = filtered_data(data, report_from, report_to)
        st.metric("Entries in period", len(report_data))

        report = make_report(report_data, report_from, report_to, report_style)
        st.markdown(report)

        st.download_button(
            "⬇️ Download Markdown Report",
            report,
            file_name=f"work_report_{report_from}_{report_to}.md",
            mime="text/markdown",
            use_container_width=True,
        )

        if not report_data.empty:
            csv = report_data.to_csv(index=False)
            st.download_button(
                "⬇️ Download Period CSV",
                csv,
                file_name=f"work_archive_{report_from}_{report_to}.csv",
                mime="text/csv",
                use_container_width=True,
            )


# ARCHIVE
if section == NAV_ARCHIVE:
    st.subheader("Archive overview")

    if data.empty:
        st.info("The archive is empty.")
    else:
        days_recorded = data["Date"].nunique()
        total_entries = len(data)
        first_day = min(data["Date"])
        last_day = max(data["Date"])

        a1, a2, a3 = st.columns(3)
        a1.metric("Days recorded", days_recorded)
        a2.metric("Total entries", total_entries)
        a3.metric("Archive span", f"{first_day} → {last_day}")

        st.dataframe(
            data.sort_values(["Date", "Category", "Field"], ascending=[False, True, True]),
            use_container_width=True,
            hide_index=True,
        )

        st.download_button(
            "⬇️ Download Excel Archive",
            excel_bytes(data),
            file_name="daily_work_archive.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
